"""
Excel utilities for reading and extracting data from Excel files.
"""

from openpyxl import load_workbook
from typing import Dict, List, Tuple, Any
from openpyxl.utils import get_column_letter


def extract_cells_by_column_range(
    file_path: str,
    sheet_name: str,
    start_column: int,
    end_column: int,
    start_row: int = 1,
    end_row: int = None,
    include_empty: bool = False
) -> Tuple[List[List[Any]], Dict[str, Tuple[int, int]]]:
    """
    Extract cells from a specific tab in an Excel file given a column range.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet/tab to extract from
        start_column: Starting column number (1-indexed, e.g., 1 for 'A', 2 for 'B')
        end_column: Ending column number (1-indexed, inclusive)
        start_row: Starting row number (1-indexed, default: 1)
        end_row: Ending row number (1-indexed, inclusive). If None, reads to last row with data
        include_empty: Whether to include completely empty rows

    Returns:
        A tuple containing:
        - List of lists: The extracted cell values in a 2D array format
        - Dictionary mapping cell addresses to their (row, col) indices in the extracted data
          Example: {'A1': (0, 0), 'B1': (0, 1), ...}

    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the sheet doesn't exist or invalid column range

    Example:
        >>> data, mapping = extract_cells_by_column_range(
        ...     'data.xlsx', 'Sheet1', 1, 3, 1, 10
        ... )
        >>> print(data)  # 2D array of values
        [['Header1', 'Header2', 'Header3'], ['val1', 'val2', 'val3'], ...]
        >>> print(mapping)  # Cell location mapping
        {'A1': (0, 0), 'B1': (0, 1), 'C1': (0, 2), ...}
    """
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")

    worksheet = workbook[sheet_name]

    # Validate column range
    if start_column < 1 or end_column < start_column:
        raise ValueError(f"Invalid column range: start_column={start_column}, end_column={end_column}")

    # Determine end_row if not provided
    if end_row is None:
        end_row = worksheet.max_row

    extracted_data = []
    cell_mapping = {}

    data_row_index = 0
    for row_idx in range(start_row, end_row + 1):
        row_data = []
        row_has_data = False

        for col_idx in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            row_data.append(cell_value)

            # Track cell location in the mapping
            cell_address = f"{get_column_letter(col_idx)}{row_idx}"
            cell_mapping[cell_address] = (data_row_index, col_idx - start_column)

            if cell_value is not None:
                row_has_data = True

        # Add row if it has data or if include_empty is True
        if row_has_data or include_empty:
            extracted_data.append(row_data)
            data_row_index += 1

    workbook.close()

    return extracted_data, cell_mapping


def extract_cells_by_column_letters(
    file_path: str,
    sheet_name: str,
    start_column: str,
    end_column: str,
    start_row: int = 1,
    end_row: int = None,
    include_empty: bool = False
) -> Tuple[List[List[Any]], Dict[str, Tuple[int, int]]]:
    """
    Extract cells from a specific tab in an Excel file given a column range using column letters.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet/tab to extract from
        start_column: Starting column letter (e.g., 'A', 'AB')
        end_column: Ending column letter (e.g., 'C', 'Z'), inclusive
        start_row: Starting row number (1-indexed, default: 1)
        end_row: Ending row number (1-indexed, inclusive). If None, reads to last row with data
        include_empty: Whether to include completely empty rows

    Returns:
        A tuple containing:
        - List of lists: The extracted cell values in a 2D array format
        - Dictionary mapping cell addresses to their (row, col) indices in the extracted data

    Raises:
        ValueError: If column letters are invalid or the sheet doesn't exist

    Example:
        >>> data, mapping = extract_cells_by_column_letters(
        ...     'data.xlsx', 'Sheet1', 'A', 'C', 1, 10
        ... )
    """
    # Convert column letters to numbers
    from openpyxl.utils import column_index_from_string

    try:
        start_col_num = column_index_from_string(start_column)
        end_col_num = column_index_from_string(end_column)
    except ValueError as e:
        raise ValueError(f"Invalid column letters: {e}")

    return extract_cells_by_column_range(
        file_path, sheet_name, start_col_num, end_col_num, start_row, end_row, include_empty
    )
